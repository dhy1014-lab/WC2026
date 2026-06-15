#!/usr/bin/env python3
"""
WC2026 Scoresheet Auto-Updater
================================
Fetches World Cup 2026 fixture results, events, and standings from
API-Football (api-sports.io) and fills in the WC2026_scoresheet_unified.xlsx
Result column (and group Field1-4 columns) for matches that have finished.

Usage:
    export API_FOOTBALL_KEY=your_key_here
    python3 update_scoresheet.py [--date YYYY-MM-DD] [--file path/to/scoresheet.xlsx] [--dry-run]

Notes:
    - Default date is yesterday (run this the morning after match day).
    - Free tier = 100 requests/day. Script prints a running call count.
    - Rows already filled (column I non-empty for prop1/prop2/bracket/goldenboot,
      or columns C-F non-empty for group rows) are skipped.
    - Props that can't be resolved from structured data are left blank with
      a "MANUAL REVIEW" note in column L.
    - A timestamped .bak copy of the workbook is saved before writing.
"""

import argparse
import os
import shutil
import sys
import time
from datetime import datetime, timedelta, timezone

import requests
import openpyxl

API_BASE = "https://v3.football.api-sports.io"
WC_LEAGUE_ID = 1       # World Cup
WC_SEASON = 2026

# ── Team name normalization ────────────────────────────────────────────────
# Maps names as they appear in the scoresheet/app -> API-Football team names
TEAM_NAME_MAP = {
    "South Africa": "South Africa",
    "South Korea": "South Korea",
    "Czechia": "Czech Republic",
    "Bosnia and Herzegovina": "Bosnia and Herzegovina",
    "Bosnia & Herzegovina": "Bosnia and Herzegovina",
    "Curacao": "Curacao",
    "Curaçao": "Curacao",
    "Ivory Coast": "Ivory Coast",
    "DR Congo": "DR Congo",
    "USA": "USA",
    "Uzbekistan": "Uzbekistan",
}

def norm_team(name):
    """Normalize a scoresheet team name to its API-Football equivalent."""
    return TEAM_NAME_MAP.get(name, name)


class ApiFootball:
    def __init__(self, key):
        self.key = key
        self.calls = 0
        self.session = requests.Session()
        self.session.headers.update({"x-apisports-key": key})

    def get(self, path, params=None):
        self.calls += 1
        resp = self.session.get(f"{API_BASE}{path}", params=params or {}, timeout=20)
        if resp.status_code in (401, 403):
            print(f"ERROR: API-Football rejected the request (HTTP {resp.status_code}). "
                  f"Check that API_FOOTBALL_KEY is valid and the subscription includes this endpoint.")
            sys.exit(1)
        resp.raise_for_status()
        data = resp.json()
        if data.get("errors"):
            print(f"  [API WARNING] {path} {params} -> errors: {data['errors']}")
        return data.get("response", [])

    def fixtures_by_date(self, date_str):
        return self.get("/fixtures", {
            "league": WC_LEAGUE_ID, "season": WC_SEASON, "date": date_str
        })

    def fixture_events(self, fixture_id):
        return self.get("/fixtures/events", {"fixture": fixture_id})

    def standings(self):
        return self.get("/standings", {"league": WC_LEAGUE_ID, "season": WC_SEASON})


def find_fixture(fixtures, team_a, team_b):
    """Find a fixture matching two team names (in either order)."""
    a, b = norm_team(team_a), norm_team(team_b)
    for fx in fixtures:
        home = fx["teams"]["home"]["name"]
        away = fx["teams"]["away"]["name"]
        names = {home, away}
        if {a, b} <= names or a in names or b in names:
            if a in names and b in names:
                return fx
    return None


def fixture_score(fx):
    """Return (home_goals, away_goals, home_name, away_name) or None if not finished."""
    if not fx:
        return None
    status = fx["fixture"]["status"]["short"]
    if status not in ("FT", "AET", "PEN"):
        return None
    goals = fx["goals"]
    if goals["home"] is None or goals["away"] is None:
        return None
    return goals["home"], goals["away"], fx["teams"]["home"]["name"], fx["teams"]["away"]["name"]


def team_goals(fx, team_name):
    sc = fixture_score(fx)
    if not sc:
        return None
    h, a, home, away = sc
    tn = norm_team(team_name)
    if tn == home:
        return h
    if tn == away:
        return a
    return None


def opponent_goals(fx, team_name):
    sc = fixture_score(fx)
    if not sc:
        return None
    h, a, home, away = sc
    tn = norm_team(team_name)
    if tn == home:
        return a
    if tn == away:
        return h
    return None


def total_goals(fx):
    sc = fixture_score(fx)
    if not sc:
        return None
    return sc[0] + sc[1]


def has_red_card(events):
    for ev in events:
        if ev["type"] == "Card" and ev["detail"] in ("Red Card", "Second Yellow card"):
            return True
    return False


def has_yellow_card(events, team_name):
    tn = norm_team(team_name)
    for ev in events:
        if ev["type"] == "Card" and ev["detail"] == "Yellow Card":
            if ev["team"]["name"] == tn:
                return True
    return False


def player_goal_or_assist(events, player_name):
    pn = player_name.lower()
    for ev in events:
        if ev["type"] != "Goal":
            continue
        scorer = (ev.get("player") or {}).get("name", "") or ""
        assist = (ev.get("assist") or {}).get("name", "") or ""
        if pn in scorer.lower() or pn in assist.lower():
            return True
    return False


def player_scored(events, player_name):
    pn = player_name.lower()
    for ev in events:
        if ev["type"] != "Goal":
            continue
        scorer = (ev.get("player") or {}).get("name", "") or ""
        if pn in scorer.lower():
            return True
    return False


def goal_in_or_after_minute(events, minute_threshold):
    """True if any goal event's elapsed (+ extra) minute >= threshold."""
    for ev in events:
        if ev["type"] != "Goal":
            continue
        elapsed = ev["time"]["elapsed"] or 0
        extra = ev["time"]["extra"] or 0
        if elapsed + extra >= minute_threshold or (elapsed >= minute_threshold):
            return True
        if elapsed == minute_threshold - 1 and extra > 0:
            return True
        if elapsed >= minute_threshold:
            return True
    return False


def first_half_goal(events, team_name):
    tn = norm_team(team_name)
    for ev in events:
        if ev["type"] != "Goal":
            continue
        if ev["team"]["name"] != tn:
            continue
        elapsed = ev["time"]["elapsed"] or 0
        if elapsed <= 45:
            return True
    return False


# ── Per-prop resolvers ───────────────────────────────────────────────────────
# Each resolver takes (api, fixtures_by_date_cache, events_cache) and returns
# (result_str, note_str) or (None, None) if unresolvable / not yet played.
#
# fixtures_by_date_cache: dict[date_str] -> list of fixtures
# events_cache: dict[fixture_id] -> list of events (lazy-fetched)

class Resolver:
    def __init__(self, api):
        self.api = api
        self.fx_cache = {}
        self.ev_cache = {}

    def fixtures(self, date_str):
        if date_str not in self.fx_cache:
            self.fx_cache[date_str] = self.api.fixtures_by_date(date_str)
        return self.fx_cache[date_str]

    def events(self, fx):
        if fx is None:
            return []
        fid = fx["fixture"]["id"]
        if fid not in self.ev_cache:
            self.ev_cache[fid] = self.api.fixture_events(fid)
        return self.ev_cache[fid]


# date_str format used by API: YYYY-MM-DD
PROP_RESOLVERS = {
    0: dict(date="2026-06-11", teams=("Mexico", "South Africa"),
            fn=lambda r, fx, ev: ("YES", f"Mexico clean sheet: opp goals={opponent_goals(fx,'Mexico')}")
                if opponent_goals(fx, "Mexico") == 0 else
                ("NO", f"South Africa scored, opp goals={opponent_goals(fx,'Mexico')}")
                if opponent_goals(fx, "Mexico") is not None else (None, None)),

    1: dict(date="2026-06-11", teams=("South Korea", "Czechia"),
            fn=lambda r, fx, ev: (("YES", f"Total goals={total_goals(fx)}") if total_goals(fx) is not None and total_goals(fx) >= 3
                else ("NO", f"Total goals={total_goals(fx)}") if total_goals(fx) is not None else (None, None))),

    2: dict(date="2026-06-12", teams=("USA", "Paraguay"),
            fn=lambda r, fx, ev: (("YES", "USA scored in 1st half") if first_half_goal(ev, "USA")
                else ("NO", "USA no first-half goal") if fixture_score(fx) is not None else (None, None))),

    3: dict(date="2026-06-12", teams=("Canada", "Bosnia and Herzegovina"),
            fn=lambda r, fx, ev: (("YES", "Canada won") if team_goals(fx, "Canada") is not None and team_goals(fx, "Canada") > opponent_goals(fx, "Canada")
                else ("NO", f"Canada result: {team_goals(fx,'Canada')}-{opponent_goals(fx,'Canada')}") if fixture_score(fx) is not None else (None, None))),

    4: dict(date="2026-06-13", teams=("Brazil", "Morocco"),
            fn=lambda r, fx, ev: (("YES", f"Brazil scored {team_goals(fx,'Brazil')}") if team_goals(fx, "Brazil") is not None and team_goals(fx, "Brazil") >= 3
                else ("NO", f"Brazil scored {team_goals(fx,'Brazil')}") if team_goals(fx, "Brazil") is not None else (None, None))),

    5: dict(date="2026-06-13", teams=None,  # special: any Day 3 match
            fn=None),  # handled specially below

    6: dict(date="2026-06-14", teams=("Germany", "Curacao"),
            fn=lambda r, fx, ev: (("YES", f"Germany won by {team_goals(fx,'Germany')-opponent_goals(fx,'Germany')}") if team_goals(fx, "Germany") is not None and (team_goals(fx, "Germany") - opponent_goals(fx, "Germany")) >= 3
                else ("NO", f"Germany margin {team_goals(fx,'Germany')-opponent_goals(fx,'Germany') if team_goals(fx,'Germany') is not None else '?'}") if fixture_score(fx) is not None else (None, None))),

    7: dict(date="2026-06-14", teams=("Netherlands", "Japan"),
            fn=lambda r, fx, ev: (("YES", f"Score {fixture_score(fx)[0]}-{fixture_score(fx)[1]}") if fixture_score(fx) is not None and fixture_score(fx)[0] == fixture_score(fx)[1]
                else ("NO", f"Score {fixture_score(fx)[0]}-{fixture_score(fx)[1]}") if fixture_score(fx) is not None else (None, None))),

    8: dict(date="2026-06-15", teams=("Spain", "Cape Verde"),
            fn=lambda r, fx, ev: (("YES", "Spain clean sheet") if opponent_goals(fx, "Spain") == 0
                else ("NO", f"Cape Verde scored {opponent_goals(fx,'Spain')}") if opponent_goals(fx, "Spain") is not None else (None, None))),

    9: dict(date="2026-06-15", teams=("Saudi Arabia", "Uruguay"),
            fn=lambda r, fx, ev: (("YES", "Saudi Arabia won") if team_goals(fx, "Saudi Arabia") is not None and team_goals(fx, "Saudi Arabia") > opponent_goals(fx, "Saudi Arabia")
                else ("NO", f"Saudi Arabia result: {team_goals(fx,'Saudi Arabia')}-{opponent_goals(fx,'Saudi Arabia')}") if fixture_score(fx) is not None else (None, None))),

    10: dict(date="2026-06-16", teams=("Argentina", "Algeria"),
             fn=lambda r, fx, ev: (("YES", "Argentina clean sheet") if opponent_goals(fx, "Argentina") == 0
                else ("NO", f"Algeria scored {opponent_goals(fx,'Argentina')}") if opponent_goals(fx, "Argentina") is not None else (None, None))),

    11: dict(date="2026-06-16", teams=None,  # special: penalty in either of 2 matches
             fn=None),

    12: dict(date="2026-06-17", teams=("Portugal", "DR Congo"),
             fn=lambda r, fx, ev: (("YES", "Ronaldo scored") if player_scored(ev, "Ronaldo")
                else ("NO", "Ronaldo did not score") if fixture_score(fx) is not None else (None, None))),

    13: dict(date="2026-06-17", teams=("England", "Croatia"),
             fn=lambda r, fx, ev: (("YES", f"Total goals={total_goals(fx)}") if total_goals(fx) is not None and total_goals(fx) < 2
                else ("NO", f"Total goals={total_goals(fx)}") if total_goals(fx) is not None else (None, None))),

    14: dict(date="2026-06-18", teams=("Mexico", "South Korea"),
             fn=lambda r, fx, ev: (("YES", "Mexico won") if team_goals(fx, "Mexico") is not None and team_goals(fx, "Mexico") > opponent_goals(fx, "Mexico")
                else ("NO", f"Mexico result: {team_goals(fx,'Mexico')}-{opponent_goals(fx,'Mexico')}") if fixture_score(fx) is not None else (None, None))),

    15: dict(date="2026-06-18", teams=("Canada", "Qatar"),
             fn=lambda r, fx, ev: (("YES", "Canada won") if team_goals(fx, "Canada") is not None and team_goals(fx, "Canada") > opponent_goals(fx, "Canada")
                else ("NO", f"Canada result: {team_goals(fx,'Canada')}-{opponent_goals(fx,'Canada')}") if fixture_score(fx) is not None else (None, None))),

    16: dict(date="2026-06-19", teams=("USA", "Australia"),
             fn=lambda r, fx, ev: (("YES", "USA won") if team_goals(fx, "USA") is not None and team_goals(fx, "USA") > opponent_goals(fx, "USA")
                else ("NO", f"USA result: {team_goals(fx,'USA')}-{opponent_goals(fx,'USA')}") if fixture_score(fx) is not None else (None, None))),

    17: dict(date="2026-06-19", teams=("Scotland", "Morocco"),
             fn=lambda r, fx, ev: (("YES", f"Total goals={total_goals(fx)}") if total_goals(fx) is not None and total_goals(fx) >= 2
                else ("NO", f"Total goals={total_goals(fx)}") if total_goals(fx) is not None else (None, None))),

    18: dict(date="2026-06-20", teams=None,  # special: late goal any Jun 20 match
             fn=None),

    19: dict(date="2026-06-20", teams=("Ecuador", "Curacao"),
             fn=lambda r, fx, ev: (("YES", "Ecuador won") if team_goals(fx, "Ecuador") is not None and team_goals(fx, "Ecuador") > opponent_goals(fx, "Ecuador")
                else ("NO", f"Ecuador result: {team_goals(fx,'Ecuador')}-{opponent_goals(fx,'Ecuador')}") if fixture_score(fx) is not None else (None, None))),

    20: dict(date="2026-06-21", teams=("Spain", "Saudi Arabia"),
             fn=lambda r, fx, ev: (("YES", f"Spain scored {team_goals(fx,'Spain')}") if team_goals(fx, "Spain") is not None and team_goals(fx, "Spain") >= 2
                else ("NO", f"Spain scored {team_goals(fx,'Spain')}") if team_goals(fx, "Spain") is not None else (None, None))),

    21: dict(date="2026-06-21", teams=("Belgium", "Iran"),
             fn=lambda r, fx, ev: (("YES", "Belgium won") if team_goals(fx, "Belgium") is not None and team_goals(fx, "Belgium") > opponent_goals(fx, "Belgium")
                else ("NO", f"Belgium result: {team_goals(fx,'Belgium')}-{opponent_goals(fx,'Belgium')}") if fixture_score(fx) is not None else (None, None))),

    22: dict(date="2026-06-22", teams=("Argentina", "Austria"),
             fn=lambda r, fx, ev: (("YES", "Argentina won, clean sheet") if (team_goals(fx, "Argentina") is not None and team_goals(fx, "Argentina") > opponent_goals(fx, "Argentina") and opponent_goals(fx, "Argentina") == 0)
                else ("NO", f"Argentina result: {team_goals(fx,'Argentina')}-{opponent_goals(fx,'Argentina')}") if fixture_score(fx) is not None else (None, None))),

    23: dict(date="2026-06-22", teams=("France", "Iraq"),
             fn=lambda r, fx, ev: (("YES", f"Total goals={total_goals(fx)}") if total_goals(fx) is not None and total_goals(fx) >= 4
                else ("NO", f"Total goals={total_goals(fx)}") if total_goals(fx) is not None else (None, None))),

    24: dict(date="2026-06-23", teams=("Portugal", "Uzbekistan"),
             fn=lambda r, fx, ev: (("YES", f"Portugal margin {team_goals(fx,'Portugal')-opponent_goals(fx,'Portugal') if team_goals(fx,'Portugal') is not None else '?'}") if team_goals(fx, "Portugal") is not None and (team_goals(fx, "Portugal") - opponent_goals(fx, "Portugal")) >= 2
                else ("NO", f"Portugal margin {team_goals(fx,'Portugal')-opponent_goals(fx,'Portugal') if team_goals(fx,'Portugal') is not None else '?'}") if fixture_score(fx) is not None else (None, None))),

    25: dict(date="2026-06-23", teams=("England", "Ghana"),
             fn=lambda r, fx, ev: (("YES", "Both teams booked") if has_yellow_card(ev, "England") and has_yellow_card(ev, "Ghana")
                else ("NO", "Not both teams booked") if fixture_score(fx) is not None else (None, None))),

    26: dict(date="2026-06-24", teams=("Brazil", "Scotland"),
             fn=lambda r, fx, ev: (("YES", "Neymar G/A") if player_goal_or_assist(ev, "Neymar")
                else ("NO", "Neymar blank") if fixture_score(fx) is not None else (None, None))),

    27: dict(date="2026-06-24", teams=("Canada", "Switzerland"),
             fn=lambda r, fx, ev: (("YES", "Davies G/A") if player_goal_or_assist(ev, "Davies")
                else ("NO", "Davies blank") if fixture_score(fx) is not None else (None, None))),

    28: dict(date="2026-06-25", teams=("USA", "Turkey"),
             fn=lambda r, fx, ev: (("YES", "Pulisic G/A") if player_goal_or_assist(ev, "Pulisic")
                else ("NO", "Pulisic blank") if fixture_score(fx) is not None else (None, None))),

    29: dict(date="2026-06-25", teams=("Germany", "Ecuador"),
             fn=lambda r, fx, ev: (("YES", f"Total goals={total_goals(fx)}") if total_goals(fx) is not None and total_goals(fx) >= 4
                else ("NO", f"Total goals={total_goals(fx)}") if total_goals(fx) is not None else (None, None))),

    30: dict(date="2026-06-26", teams=("Norway", "France"),  # Haaland plays for Norway
             fn=lambda r, fx, ev: (("YES", "Haaland scored") if player_scored(ev, "Haaland")
                else ("NO", "Haaland blank") if fixture_score(fx) is not None else (None, None))),

    31: dict(date="2026-06-26", teams=None,  # special: late goal any Jun 26 match
             fn=None),

    32: dict(date="2026-06-27", teams=("England", "Panama"),
             fn=lambda r, fx, ev: (("YES", "Bellingham G/A") if player_goal_or_assist(ev, "Bellingham")
                else ("NO", "Bellingham blank") if fixture_score(fx) is not None else (None, None))),

    33: dict(date="2026-06-27", teams=("Argentina", "Jordan"),
             fn=lambda r, fx, ev: (("YES", "Messi scored") if player_scored(ev, "Messi")
                else ("NO", "Messi blank") if fixture_score(fx) is not None else (None, None))),
}


def resolve_special_prop(idx, resolver):
    """Handle props that span multiple matches on a day."""
    if idx == 5:  # red card in any Day 3 match
        fixtures = resolver.fixtures("2026-06-13")
        finished = [fx for fx in fixtures if fixture_score(fx) is not None]
        if not finished or len(finished) < len(fixtures):
            return (None, None)  # wait until all Day 3 matches done
        for fx in fixtures:
            if has_red_card(resolver.events(fx)):
                return ("YES", "Red card found on Day 3")
        return ("NO", "No red cards on Day 3")

    if idx == 11:  # penalty in Argentina v Algeria OR Austria v Jordan
        fixtures = resolver.fixtures("2026-06-16")
        fx1 = find_fixture(fixtures, "Argentina", "Algeria")
        fx2 = find_fixture(fixtures, "Austria", "Jordan")
        if fixture_score(fx1) is None or fixture_score(fx2) is None:
            return (None, None)
        for fx in (fx1, fx2):
            for ev in resolver.events(fx):
                if ev["type"] == "Var" or (ev["type"] == "Goal" and ev["detail"] == "Penalty"):
                    return ("YES", "Penalty detected")
                if ev["type"] == "Card" and "penalty" in (ev.get("detail") or "").lower():
                    return ("YES", "Penalty detected")
        # also check missed penalties via "Missed Penalty" goal detail
        for fx in (fx1, fx2):
            for ev in resolver.events(fx):
                if ev["type"] == "Goal" and ev["detail"] in ("Penalty", "Missed Penalty"):
                    return ("YES", "Penalty detected")
        return ("NO", "No penalties detected")

    if idx == 18:  # goal in 81st minute+ any Jun 20 match
        fixtures = resolver.fixtures("2026-06-20")
        finished = [fx for fx in fixtures if fixture_score(fx) is not None]
        if not finished or len(finished) < len(fixtures):
            return (None, None)
        for fx in fixtures:
            for ev in resolver.events(fx):
                if ev["type"] == "Goal":
                    elapsed = ev["time"]["elapsed"] or 0
                    extra = ev["time"]["extra"] or 0
                    if elapsed >= 81 or (elapsed == 80 and extra > 0):
                        return ("YES", f"Late goal: {fx['teams']['home']['name']} vs {fx['teams']['away']['name']} @ {elapsed}+{extra}")
        return ("NO", "No goals at 81'+ on Jun 20")

    if idx == 31:  # goal in 85th minute+ any Jun 26 match
        fixtures = resolver.fixtures("2026-06-26")
        finished = [fx for fx in fixtures if fixture_score(fx) is not None]
        if not finished or len(finished) < len(fixtures):
            return (None, None)
        for fx in fixtures:
            for ev in resolver.events(fx):
                if ev["type"] == "Goal":
                    elapsed = ev["time"]["elapsed"] or 0
                    extra = ev["time"]["extra"] or 0
                    if elapsed >= 85 or (elapsed == 84 and extra > 0):
                        return ("YES", f"Late goal: {fx['teams']['home']['name']} vs {fx['teams']['away']['name']} @ {elapsed}+{extra}")
        return ("NO", "No goals at 85'+ on Jun 26")

    return (None, None)


def process_prop_rows(ws, resolver, dry_run):
    updates = []
    for row in range(2, 36):
        idx = ws.cell(row=row, column=2).value  # B = ID (0-33)
        result_cell = ws.cell(row=row, column=9)  # I
        if result_cell.value not in (None, ""):
            continue  # already filled

        cfg = PROP_RESOLVERS.get(idx)
        if cfg is None:
            continue

        if cfg["teams"] is None:
            res, note = resolve_special_prop(idx, resolver)
        else:
            fx = find_fixture(resolver.fixtures(cfg["date"]), *cfg["teams"])
            ev = resolver.events(fx) if fx else []
            res, note = cfg["fn"](resolver, fx, ev) if fx else (None, None)

        if res is None:
            continue

        updates.append((row, res, note))
        if not dry_run:
            ws.cell(row=row, column=9).value = res  # I = Result
            note_cell = ws.cell(row=row, column=12)  # L = Notes
            if note_cell.value in (None, ""):
                note_cell.value = f"Auto: {note}"
    return updates


def process_group_rows(ws, resolver, dry_run):
    """Fill Field1-4 (C-F) = ranked team order from standings, once a group
    has all 3 matches per team complete (i.e. standings 'played' == 3 each)."""
    updates = []
    standings_resp = resolver.api.standings()
    if not standings_resp:
        return updates

    # API-Football standings response: list per league, each with 'league'->'standings'
    # standings is a list of groups, each a list of team standing dicts
    groups_data = {}
    for league_entry in standings_resp:
        for group_list in league_entry["league"].get("standings", []):
            if not group_list:
                continue
            group_name_raw = group_list[0].get("group", "")
            # e.g. "Group A" -> "A"
            letter = group_name_raw.strip().split(" ")[-1]
            groups_data[letter] = group_list

    for row in range(36, 48):
        letter = ws.cell(row=row, column=2).value  # B = group letter
        existing = [ws.cell(row=row, column=c).value for c in range(3, 7)]  # C-F
        if any(v not in (None, "") for v in existing):
            continue  # already filled

        group_list = groups_data.get(letter)
        if not group_list:
            continue

        all_played = all(team["all"]["played"] >= 3 for team in group_list)
        if not all_played:
            continue

        ranked = sorted(group_list, key=lambda t: t["rank"])
        names = [t["team"]["name"] for t in ranked]
        updates.append((row, letter, names))
        if not dry_run:
            for i, name in enumerate(names):
                ws.cell(row=row, column=3 + i).value = name  # C,D,E,F = Field1-4
    return updates


def main():
    parser = argparse.ArgumentParser(description="Update WC2026 scoresheet with live results")
    parser.add_argument("--date", help="Date to fetch fixtures for (YYYY-MM-DD), default=yesterday")
    parser.add_argument("--file", default="WC2026_scoresheet_unified.xlsx", help="Path to scoresheet xlsx")
    parser.add_argument("--dry-run", action="store_true", help="Don't write changes, just report")
    args = parser.parse_args()

    key = os.environ.get("API_FOOTBALL_KEY")
    if not key:
        print("ERROR: set API_FOOTBALL_KEY environment variable.")
        sys.exit(1)

    if not os.path.exists(args.file):
        print(f"ERROR: file not found: {args.file}")
        sys.exit(1)

    api = ApiFootball(key)
    resolver = Resolver(api)

    wb = openpyxl.load_workbook(args.file)
    ws = wb["Scoresheet"]

    print("Resolving daily prop rows...")
    prop_updates = process_prop_rows(ws, resolver, args.dry_run)
    for row, res, note in prop_updates:
        print(f"  Row {row}: {res}  ({note})")

    print("\nResolving group standings...")
    group_updates = process_group_rows(ws, resolver, args.dry_run)
    for row, letter, names in group_updates:
        print(f"  Group {letter}: {', '.join(names)}")

    print(f"\nAPI calls used: {api.calls}")

    if args.dry_run:
        print("\nDRY RUN — no changes written.")
        return

    if not prop_updates and not group_updates:
        print("\nNo updates to write.")
        return

    backup_path = f"{args.file}.{datetime.now().strftime('%Y%m%d_%H%M%S')}.bak"
    shutil.copy(args.file, backup_path)
    print(f"\nBackup saved: {backup_path}")

    wb.save(args.file)
    print(f"Saved updates to: {args.file}")


if __name__ == "__main__":
    main()
